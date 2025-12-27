import os
import json
import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from openpyxl import load_workbook

TOKEN = os.getenv("BOT_TOKEN")

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)

MAN_FILE = "managers.json"
ADDR_FILE = "addresses.json"
EXCEL_FILE = "data.xlsx"

user_state = {}

ZONES = {
    1: ["героїв дніпра","мінська","оболонь","почайна","тарса шевченка",
        "контрактова площа","поштова площа","майдан незалежності",
        "площа українських героїв","олімпійська","палац україна",
        "либідська","академмістечко","житомирська","святошин","нивки",
        "берестейська","шулявська","політехнічний інститут",
        "вокзальна","університет","театральна","хрещатик",
        "арсенальна","дорогожичі","печерськ","сирець"],

    2: ["звіринецька","деміївська","голосіївська",
        "васильківська","вднх","іподром","теремки"],

    3: ["дніпро","гідропарк","лівобережна",
        "дарниця","чернігівська","лісова","троєщина"],

    4: ["славутич","осокорки","позняки",
        "харківська","вирлиця",
        "бориспільська","червоний хутір"]
}

def load_json(file, default):
    if not os.path.exists(file):
        return default
    with open(file, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(file, data):
    with open(file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def is_manager(user_id):
    return str(user_id) in load_json(MAN_FILE, [])

def get_zone(station):
    station = station.lower().strip()
    for z, sts in ZONES.items():
        if station in sts:
            return z
    return None

def find_person(surname):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == surname.lower():
            return {
                "surname": row[0],
                "address": row[1],
                "station": row[2],
                "zone": get_zone(row[2])
            }
    return None

@dp.message_handler(commands=["start","info"])
async def info(msg: types.Message):
    await msg.answer(
        "Бот працює\n\n"
        "/add — додати пасажира\n"
        "/del — видалити пасажира\n"
        "/list — список по зонах\n"
        "/add_Man — додати менеджера\n"
        "/del_Man — видалити менеджера"
    )

@dp.message_handler(commands=["add"])
async def add_cmd(msg: types.Message):
    if not is_manager(msg.from_user.id):
        return await msg.answer("❌ Ти не менеджер")
    user_state[msg.from_user.id] = "wait_surname"
    await msg.answer("Введи **прізвище** наступним повідомленням")

@dp.message_handler()
async def handle_text(msg: types.Message):
    uid = msg.from_user.id
    if user_state.get(uid) != "wait_surname":
        return

    person = find_person(msg.text.strip())
    if not person:
        return await msg.answer("❌ Прізвище не знайдено в Excel")

    data = load_json(ADDR_FILE, [])
    data.append(person)
    save_json(ADDR_FILE, data)

    user_state.pop(uid)
    await msg.answer(
        f"✅ Додано:\n{person['surname']}\n"
        f"{person['address']}\n"
        f"Метро: {person['station']}\n"
        f"Зона: {person['zone']}"
    )

@dp.message_handler(commands=["list"])
async def list_cmd(msg: types.Message):
    data = load_json(ADDR_FILE, [])
    if not data:
        return await msg.answer("Список порожній")

    text = ""
    for z in range(1,5):
        group = [d for d in data if d["zone"] == z]
        if group:
            text += f"\n🚕 Зона {z}:\n"
            for p in group:
                text += f"- {p['surname']} | {p['address']}\n"
    await msg.answer(text)

@dp.message_handler(commands=["del"])
async def del_cmd(msg: types.Message):
    if not is_manager(msg.from_user.id):
        return
    user_state[msg.from_user.id] = "del"
    await msg.answer("Введи прізвище для видалення")

@dp.message_handler(commands=["add_Man"])
async def add_man(msg: types.Message):
    ids = load_json(MAN_FILE, [])
    parts = msg.text.split()
    if len(parts) != 2:
        return await msg.answer("/add_Man ID")
    ids.append(parts[1])
    save_json(MAN_FILE, ids)
    await msg.answer("✅ Менеджер доданий")

@dp.message_handler(commands=["del_Man"])
async def del_man(msg: types.Message):
    ids = load_json(MAN_FILE, [])
    parts = msg.text.split()
    if len(parts) != 2:
        return await msg.answer("/del_Man ID")
    if parts[1] in ids:
        ids.remove(parts[1])
        save_json(MAN_FILE, ids)
    await msg.answer("✅ Менеджер видалений")

if __name__ == "__main__":
    executor.start_polling(dp)
